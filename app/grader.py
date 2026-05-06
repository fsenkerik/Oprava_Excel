import os
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import openpyxl

GRADE_THRESHOLDS = [
    (90, 1),
    (75, 2),
    (60, 3),
    (40, 4),
    (0, 5),
]

GRADE_LABELS = {
    1: "Výborný",
    2: "Chvalitebný",
    3: "Dobrý",
    4: "Dostatečný",
    5: "Nedostatečný",
}


def compute_grade(score: int, max_score: int) -> int:
    if max_score == 0:
        return 5
    pct = (score / max_score) * 100
    for threshold, grade in GRADE_THRESHOLDS:
        if pct >= threshold:
            return grade
    return 5


def grade_label(grade: int) -> str:
    return GRADE_LABELS.get(grade, "")


def _round2(value) -> Decimal:
    try:
        return Decimal(str(float(value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _normalize_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip().lower()


def _is_formula(cell_value) -> bool:
    """Zjistí zda hodnota buňky je vzorec (začíná =)."""
    return isinstance(cell_value, str) and cell_value.startswith('=')


def _values_match(student_val, key_val, comparison: str) -> bool:
    """Porovná výslednou hodnotu studenta se správnou hodnotou z klíče."""
    if student_val is None or student_val == '':
        return False
    if key_val is None or key_val == '':
        return False

    if comparison == 'text':
        return _normalize_text(student_val) == _normalize_text(key_val)

    if comparison == 'numeric':
        s = _round2(student_val)
        k = _round2(key_val)
        if s is None or k is None:
            return False
        return s == k

    # mixed: zkusí numeric, pak text
    s = _round2(student_val)
    k = _round2(key_val)
    if s is not None and k is not None:
        return s == k
    return _normalize_text(student_val) == _normalize_text(key_val)


def _read_column(ws, col_letter: str, row_start: int, row_end: int) -> list:
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    return [ws.cell(row=r, column=col_idx).value for r in range(row_start, row_end + 1)]


def grade_submission(uploaded_path: str, exercise_config: dict, key_path: str) -> dict:
    # Načteme dvakrát:
    # - formula_wb: data_only=False → vidíme zda buňka obsahuje vzorec (=...)
    # - values_wb:  data_only=True  → vidíme vypočtený výsledek vzorce
    with open(uploaded_path, 'rb') as f:
        formula_wb = openpyxl.load_workbook(f, data_only=False)
    with open(uploaded_path, 'rb') as f:
        values_wb = openpyxl.load_workbook(f, data_only=True)
    with open(key_path, 'rb') as f:
        key_wb = openpyxl.load_workbook(f, data_only=True)

    task_scores = {}
    task_details = {}
    total_score = 0
    max_score = 0

    for task in exercise_config["tasks"]:
        tid = task["id"]
        sheet_name = task["sheet"]
        max_pts = task["max_points"]
        comparison = task["comparison"]
        requires_formula = task.get("requires_formula", False)
        answer_col2 = task.get("answer_col2")  # Druhý sloupec (Úkol 5)

        if sheet_name not in values_wb.sheetnames:
            task_scores[tid] = 0
            task_details[tid] = []
            max_score += max_pts
            continue

        formula_ws = formula_wb[sheet_name]
        values_ws = values_wb[sheet_name]
        key_ws = key_wb[task["key_sheet"]]

        # Správné hodnoty z klíče
        key_answers = _read_column(
            key_ws, task["key_col"],
            task["key_row_start"], task["key_row_end"]
        )

        # Odpovědi studenta — hodnoty i vzorce
        student_formulas = _read_column(
            formula_ws, task["answer_col"],
            task["answer_row_start"], task["answer_row_end"]
        )
        student_values = _read_column(
            values_ws, task["answer_col"],
            task["answer_row_start"], task["answer_row_end"]
        )

        # Druhý sloupec pro Úkol 5
        if answer_col2:
            student_formulas2 = _read_column(
                formula_ws, answer_col2,
                task["answer_row_start"], task["answer_row_end"]
            )
            student_values2 = _read_column(
                values_ws, answer_col2,
                task["answer_row_start"], task["answer_row_end"]
            )
        else:
            student_formulas2 = [None] * len(key_answers)
            student_values2 = [None] * len(key_answers)

        correct_count = 0
        details = []

        for i, kv in enumerate(key_answers):
            sf = student_formulas[i]
            sv = student_values[i]

            if requires_formula:
                # Musí být vzorec (začíná =)
                has_formula = _is_formula(sf)

                if answer_col2:
                    # Úkol 5: oba sloupce musí být vzorce se správným výsledkem
                    sf2 = student_formulas2[i]
                    sv2 = student_values2[i]
                    has_formula2 = _is_formula(sf2)
                    val_ok = _values_match(sv, kv, comparison)
                    val_ok2 = _values_match(sv2, kv, comparison)
                    correct = has_formula and has_formula2 and val_ok and val_ok2

                    if not has_formula or not has_formula2:
                        reason = "chybí vzorec"
                    elif not val_ok or not val_ok2:
                        reason = "špatný výsledek"
                    else:
                        reason = "správně"
                else:
                    # Standardní: vzorec + správný výsledek
                    val_ok = _values_match(sv, kv, comparison)
                    correct = has_formula and val_ok

                    if not has_formula:
                        reason = "chybí vzorec"
                    elif not val_ok:
                        reason = "špatný výsledek"
                    else:
                        reason = "správně"
            else:
                # Úkol 1: stačí správná hodnota (vzorec není povinný)
                correct = _values_match(sv, kv, comparison)
                reason = "správně" if correct else "špatná odpověď"

            if correct:
                correct_count += 1

            details.append({
                "question": i + 1,
                "student": str(sf) if sf is not None else "",
                "correct": correct,
                "reason": reason,
            })

        task_scores[tid] = correct_count
        task_details[tid] = details
        total_score += correct_count
        max_score += max_pts

    grade = compute_grade(total_score, max_score)

    return {
        "total_score": total_score,
        "max_score": max_score,
        "grade": grade,
        "grade_label": grade_label(grade),
        "task_scores": task_scores,
        "task_details": task_details,
    }
